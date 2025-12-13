import re
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

import wardrobe_2door
import wardrobe_1door
import loft
import wardrobe_2door_slide
import wardrobe_3door_slide
import kitchen

# --- Map wardrobe types to their form/calc functions and image paths ---
# image_paths can be 1 or 2 paths; both will be shown one below the other.
type_fns = {
    "1-Door Wardrobe": (
        wardrobe_1door.form_type1,
        wardrobe_1door.calc_type1,
        ["1door.PNG", "1door_2.PNG"]  # single image is fine
    ),
    "2-Door Normal": (
        wardrobe_2door.form_type1,
        wardrobe_2door.calc_type1,
        ["2door.jpg"]  # <— add second image if available
    ),
    "2-Door Sliding": (
        wardrobe_2door_slide.form_type1,
        wardrobe_2door_slide.calc_type1,
        ["2door_slide.png", "2door_slide_2.png"]
    ),
    "3-Door Sliding": (
        wardrobe_3door_slide.form_type1,
        wardrobe_3door_slide.calc_type1,
        ["2door_slide.png", "2door_slide_2.png"]
    ),
    "Wardrobe Loft ": (
        loft.form_type1,
        loft.calc_type1,
        ["loft1.png", "loft2.png"]
    ),
    "Kitchen": (
        kitchen.form_type1,
        kitchen.calc_type1,
        []  # no images for now
    ),

}

# ---------- Helpers: normalize DF -> new columns ----------
STD_COLS = [
    "SLNO",
    "Description",
    "Height",
    "Width",
    "Qty",
    "Material",
    "Long side 1",
    "Long side 2",
    "Short side 1",
    "Short side 2",
    "Groove",
]

def _dims(h, w):
    # compact dimension string (mm implied)
    try:
        return f"{round(float(h),1)} × {round(float(w),1)}"
    except Exception:
        return ""

def normalize_to_std(df):
    """
    Accepts a module DataFrame and returns a DataFrame with the standard columns.
    Handles three cases:
      1) DF already has the standard columns -> return as-is (reordered).
      2) DF is the old "item/height_mm/width_mm[/qty]" schema -> map directly.
      3) DF is the previous 9-column shape (Cut piece name/Wood/etc.) -> parse dims and map.
    """
    import pandas as pd

    def parse_dims(val: str):
        """Parse strings like '710 × 578', '710 x 578', or '710 × 578 = 3 qty'."""
        if not isinstance(val, str):
            return None
        text = val.lower()
        # normalize separators
        text = text.replace("×", "x").replace(" X ", "x")
        m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*x\s*([0-9]+(?:\.[0-9]+)?)", text)
        if not m:
            return None
        h = float(m.group(1))
        w = float(m.group(2))
        qty_match = re.search(r"=\s*([0-9]+)", text)
        qty = int(qty_match.group(1)) if qty_match else 1
        return h, w, qty

    # Case 1: already in standard column shape
    if set([c.lower() for c in df.columns]).issuperset(
        [c.lower() for c in STD_COLS]
    ):
        # Add SLNO if not present
        if "SLNO" not in df.columns and "slno" not in [c.lower() for c in df.columns]:
            df.insert(0, "SLNO", range(1, len(df) + 1))
        # reorder and return
        return df[[c for c in STD_COLS if c in df.columns]]

    lower = {c.lower(): c for c in df.columns}

    # Case 2: older numeric schema
    if all(k in lower for k in ["item", "height_mm", "width_mm"]):
        item_col = lower["item"]
        h_col = lower["height_mm"]
        w_col = lower["width_mm"]
        q_col = lower.get("qty", None)

        out_rows = []
        for idx, r in df.iterrows():
            out_rows.append({
                "SLNO": idx + 1,
                "Description": r[item_col],
                "Height": r[h_col],
                "Width": r[w_col],
                "Qty": r[q_col] if q_col else 1,
                "Material": "",
                "Long side 1": "",
                "Long side 2": "",
                "Short side 1": "",
                "Short side 2": "",
                "Groove": "",
            })
        return pd.DataFrame(out_rows, columns=STD_COLS)

    # Case 3: legacy shape with Wood column -> parse dims (even if some optional cols missing)
    lower_cols = set([c.lower() for c in df.columns])
    if "wood" in lower_cols and "cut piece name" in lower_cols:
        out_rows = []
        for idx, r in df.iterrows():
            desc = r.get("Cut piece name", r.get("cut piece name", ""))
            wood_val = r.get("Wood", r.get("wood", ""))
            col_lam = r.get("Colour laminate", r.get("colour laminate", ""))
            lam_color = r.get("Laminate Color", r.get("laminate color", ""))

            parsed = parse_dims(str(wood_val))
            if not parsed:
                parsed = parse_dims(str(col_lam))
            if parsed:
                h, w, q = parsed
            else:
                h, w, q = 0, 0, 1

            # Material preference: laminate color if present, otherwise the wood string (if it's not just dims)
            material_candidate = str(lam_color or "").strip()
            if not material_candidate:
                material_candidate = str(wood_val or "").strip()
            material = material_candidate

            out_rows.append({
                "SLNO": idx + 1,
                "Description": desc,
                "Height": h,
                "Width": w,
                "Qty": q,
                "Material": material,
                "Long side 1": r.get("Long side 1", ""),
                "Long side 2": r.get("Long side 2", ""),
                "Short side 1": r.get("Short side 1", ""),
                "Short side 2": r.get("Short side 2", ""),
                "Groove": r.get("Groove", ""),
            })
        return pd.DataFrame(out_rows, columns=STD_COLS)

    # Fallback: ensure SLNO exists and return as-is
    if "SLNO" not in df.columns:
        df.insert(0, "SLNO", range(1, len(df) + 1))
    return df

def safe_filename(name: str) -> str:
    # turn "Kitchen - Cabinet" into "kitchen_cabinet"
    base = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").lower()
    return base or "cutlist"

def create_styled_xlsx(df):
    """
    Create an XLSX file with styling - golden background for heading rows
    Returns: BytesIO buffer containing the XLSX file
    """
    output = BytesIO()
    
    # Create a new workbook and select the active sheet
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cut List')
        workbook = writer.book
        worksheet = writer.sheets['Cut List']
        
        # Define styles
        heading_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        heading_font = Font(bold=True, color='000000')
        normal_alignment = Alignment(horizontal='left', vertical='center')
        
        # Apply styles to rows (starting from row 2, as row 1 is header)
        for idx, row in df.iterrows():
            excel_row = idx + 2  # +2 because Excel is 1-indexed and row 1 is header
            
            # Check if this is a heading row (Material column is empty)
            is_heading = (str(row.get("Material", "")).strip() == "" or pd.isna(row.get("Material", ""))) and str(row.get("Description", "")).strip() != ""
            
            if is_heading:
                # Apply golden background and bold font to all cells in this row
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.fill = heading_fill
                    cell.font = heading_font
                    cell.alignment = normal_alignment
            else:
                # Apply normal alignment to non-heading rows
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.alignment = normal_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

# ----------------------------------------------------------

st.set_page_config(page_title=" Multi-Type Material Calculator", layout="wide")
st.title("🛠️ Multi-Type Cut Piece Calculator")

# --- Session state init ---
if "all_types_inputs" not in st.session_state:
    st.session_state["all_types_inputs"] = []
    st.session_state["all_types_labels"] = []
if "edit_index" not in st.session_state:
    st.session_state["edit_index"] = None

# --- Sidebar: Add/Edit/Manage ---
with st.sidebar:
    st.header("➕ Add New / Edit")

    # If editing, lock type and prefill fields
    if st.session_state["edit_index"] is not None:
        st.warning(f"✏ Editing Option {st.session_state['edit_index'] + 1}")
        type_label = st.session_state["all_types_labels"][st.session_state["edit_index"]]
        prefill = st.session_state["all_types_inputs"][st.session_state["edit_index"]]
    else:
        type_label = st.selectbox("Module Type", list(type_fns.keys()), key="sidebar_type")
        prefill = {}

    form_fn, _, _ = type_fns[type_label]

    with st.form("type_form"):
        button_label = "Update" if st.session_state["edit_index"] is not None else "Add"
        submitted, input_data = form_fn(prefill, button_label)
        if submitted:
            if st.session_state["edit_index"] is None:
                # Add new entry
                st.session_state["all_types_inputs"].append(input_data)
                st.session_state["all_types_labels"].append(type_label)
                st.success(f"✅ Added {type_label}")
            else:
                # Update existing
                idx = st.session_state["edit_index"]
                st.session_state["all_types_inputs"][idx] = input_data
                st.session_state["all_types_labels"][idx] = type_label
                st.session_state["edit_index"] = None
                st.success(f"✏️ Updated {type_label}")

    st.markdown("---")
    st.header("🗂 Manage Added Modules")
    for i, tname in enumerate(st.session_state["all_types_labels"]):
        st.write(f"{i+1}. {tname}")
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("✏ Edit", key=f"edit_{i}"):
                st.session_state["edit_index"] = i
                st.rerun()
        with col2:
            if st.button("📄 Duplicate", key=f"dup_{i}"):
                st.session_state["all_types_inputs"].append(
                    st.session_state["all_types_inputs"][i].copy()
                )
                st.session_state["all_types_labels"].append(tname)
        with col3:
            if st.button("🗑 Delete", key=f"del_{i}"):
                st.session_state["all_types_inputs"].pop(i)
                st.session_state["all_types_labels"].pop(i)
                if st.session_state["edit_index"] == i:
                    st.session_state["edit_index"] = None
                st.rerun()

# --- Main area: Outputs ---
if st.session_state["all_types_inputs"]:
    st.header("📋 All Materials by Module Type in mm")
    for i, tname in enumerate(st.session_state["all_types_labels"]):
        st.subheader(f"{tname} — Option {i+1}")
        form_fn, calc_fn, image_paths = type_fns[tname]

        # Only show cutpiece output (images removed as requested)
        cols = st.columns([1])
        with cols[0]:
            # Try to get a DataFrame from the module
            df = None
            module_obj = {
                "1-Door Wardrobe": wardrobe_1door,
                "2-Door Normal": wardrobe_2door,
                "2-Door Sliding": wardrobe_2door_slide,
                "3-Door Sliding": wardrobe_3door_slide,
                "Wardrobe Loft ": loft,
                "Kitchen": kitchen,
            }[tname]

            if hasattr(module_obj, "get_cutlist_df"):
                try:
                    raw_df = module_obj.get_cutlist_df(st.session_state["all_types_inputs"][i])
                except Exception as e:
                    raw_df = None
                    st.error(f"Error generating table: {e}")

                if raw_df is not None:
                    df = normalize_to_std(raw_df)

            # Display user input log (key-value) above cutpiece table for all types
            user_inputs = st.session_state["all_types_inputs"][i]
            st.markdown("**User Inputs:**")
            keys = list(user_inputs.keys())
            values = list(user_inputs.values())
            n = len(keys)
            cols = st.columns(4)
            for idx in range(n):
                col = cols[idx % 4]
                col.write(f"**{keys[idx]}**: {values[idx]}")

            if df is not None and not df.empty and set([c.lower() for c in STD_COLS]).issubset([c.lower() for c in df.columns]):
                # Ensure required columns exist (create if missing)
                for col in STD_COLS:
                    if col not in df.columns:
                        if col == "SLNO":
                            df.insert(0, "SLNO", range(1, len(df) + 1))
                        elif col in ["Height", "Width", "Qty"]:
                            df[col] = 0
                        else:
                            df[col] = ""

                # Ensure proper column order (preserve only our standard columns)
                df = df[[c for c in STD_COLS if c in df.columns]]

                # Derive edge bidding columns (all text) from Material
                if "Material" in df.columns:
                    col_str = df["Material"].astype(str)
                    mask_fb_bsl = col_str.str.contains("FB BSL", na=False)

                    # If FB BSL → only Long side 1 = "0.8 FB", others blank
                    if "Long side 1" in df.columns:
                        df.loc[mask_fb_bsl, "Long side 1"] = "0.8 FB"
                    for col in ["Short side 1", "Short side 2", "Long side 2"]:
                        if col in df.columns:
                            df.loc[mask_fb_bsl, col] = ""

                    # Else → Long side 1 = "Colour Bidding", others blank
                    if "Short side 1" in df.columns:
                        df.loc[~mask_fb_bsl, "Short side 1"] = ""
                    if "Short side 2" in df.columns:
                        df.loc[~mask_fb_bsl, "Short side 2"] = ""
                    if "Long side 2" in df.columns:
                        df.loc[~mask_fb_bsl, "Long side 2"] = ""
                    if "Long side 1" in df.columns:
                        df.loc[~mask_fb_bsl, "Long side 1"] = "Colour Bidding"

                # Ensure proper dtypes
                if "SLNO" in df.columns:
                    df["SLNO"] = pd.to_numeric(df["SLNO"], errors="coerce").fillna(0).astype(int)
                if "Height" in df.columns:
                    df["Height"] = pd.to_numeric(df["Height"], errors="coerce").fillna(0)
                if "Width" in df.columns:
                    df["Width"] = pd.to_numeric(df["Width"], errors="coerce").fillna(0)
                if "Qty" in df.columns:
                    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(int)
                for text_col in ["Description", "Material", "Short side 1", "Short side 2", "Long side 1", "Long side 2", "Groove"]:
                    if text_col in df.columns:
                        df[text_col] = df[text_col].where(df[text_col].notna(), "").astype(str)

                # Display editable table first
                st.subheader("✏️ Editable Cut List")
                st.info("💡 Click any cell to edit. Add/delete rows as needed. Changes will reflect in the preview below.")
                
                # Display editable dataframe 
                edited_df = st.data_editor(
                    df,
                    use_container_width=True,
                    num_rows="dynamic",
                    column_config={
                        "SLNO": st.column_config.NumberColumn("SLNO", width="small", disabled=True),
                        "Description": st.column_config.TextColumn("Description", width="large"),
                        "Height": st.column_config.NumberColumn("Height", width="small"),
                        "Width": st.column_config.NumberColumn("Width", width="small"),
                        "Qty": st.column_config.NumberColumn("Qty", width="small"),
                        "Material": st.column_config.TextColumn("Material", width="large"),
                        "Long side 1": st.column_config.TextColumn("Long side 1", width="medium"),
                        "Long side 2": st.column_config.TextColumn("Long side 2", width="medium"),
                        "Short side 1": st.column_config.TextColumn("Short side 1", width="medium"),
                        "Short side 2": st.column_config.TextColumn("Short side 2", width="medium"),
                        "Groove": st.column_config.TextColumn("Groove", width="medium"),
                    },
                    key=f"data_editor_{i}",
                )
                
                # Helper function to style heading rows
                def highlight_headings(row):
                    """Apply background color to heading rows (where Material column is empty)"""
                    # Check if this is a heading row (Material column is empty/blank)
                    is_heading = (str(row.get("Material", "")).strip() == "" or pd.isna(row.get("Material", ""))) and str(row.get("Description", "")).strip() != ""
                    if is_heading:
                        # Return yellow/golden background for heading rows
                        return ['background-color: #FFD700; font-weight: bold; color: #000000'] * len(row)
                    else:
                        return [''] * len(row)
                
                # Display styled preview (read-only, updates with edits)
                st.subheader("📊 Preview with Highlighted Headings")
                st.caption("Read-only preview - updates automatically with your edits above")
                # Create a formatted copy to drop trailing .0 for whole numbers
                formatted_df = edited_df.copy()

                def _fmt_num(val):
                    if pd.isna(val):
                        return ""
                    if isinstance(val, (int, float)):
                        return str(int(val)) if float(val).is_integer() else str(val)
                    return val

                for num_col in ["Height", "Width", "Qty"]:
                    if num_col in formatted_df.columns:
                        formatted_df[num_col] = formatted_df[num_col].apply(_fmt_num)

                styled_df = formatted_df.style.apply(highlight_headings, axis=1)
                st.dataframe(styled_df, use_container_width=True, height=400)

                # XLSX download with edited data and styling - add unique key to avoid duplicate error
                xlsx_buffer = create_styled_xlsx(edited_df)
                fname = f"{safe_filename(tname)}_cutlist.xlsx"
                st.download_button(
                    "📥 Download Styled Cut List as Excel", 
                    xlsx_buffer, 
                    fname, 
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    key=f"download_{i}"
                )
            else:
                # Fallback to legacy bullets
                mats = module_obj.calc_type1(st.session_state["all_types_inputs"][i])
                for line in mats:
                    st.write(line)

        st.markdown("---")
else:
    st.info("No Modules added yet. Use the sidebar to add one.")
